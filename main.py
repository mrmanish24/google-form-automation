from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import workbook, load_workbook
import time

service = Service(executable_path="chromedriver.exe")
driver = webdriver.Chrome(service=service)


#____________________________________________________________________________________________________
# wait for element (utility)
# WebDriverWait(driver, 10).until(
#     EC.presence_of_element_located((By.ID, 'exampleId'))
# )

# try caught (utility)
# try:
#     driver.find_element(By.ID, "name_field").send_keys(row["Name"])
# except Exception as e:
#     print(f"Error filling form for {row['Name']}: {e}")

# for col in sheet.iter_cols(min_row=1,max_row=6,min_col=1,max_col=7, values_only = True) :
#     print(col)

# for row in sheet.iter_rows(min_row=1,max_row=6,min_col=1,max_col=7, values_only = True) :
#_________________________________________________________________________________________________________

# driver code
wb = load_workbook("userdata.xlsx")
sheet = wb.active

driver.get('https://docs.google.com/forms/d/e/1FAIpQLSc3currLhMcl-dS-EuUVfMcZ-LWdDHNjKPkPLx56Qv8sq4ypw/viewform?usp=sf_link')


userName_id = '//*[@id="mG61Hd"]/div[2]/div/div[2]/div[1]/div/div/div[2]/div/div[1]/div/div[1]/input'
phrase_id = '//*[@id="mG61Hd"]/div[2]/div/div[2]/div[2]/div/div/div[2]/div/div[1]/div/div[1]/input'
animal_id = '//*[@id="mG61Hd"]/div[2]/div/div[2]/div[3]/div/div/div[2]/div/div[1]/div/div[1]/input'
talent_id = '//*[@id="mG61Hd"]/div[2]/div/div[2]/div[4]/div/div/div[2]/div/div[1]/div/div[1]/input'
planet_id = '//*[@id="mG61Hd"]/div[2]/div/div[2]/div[5]/div/div/div[2]/div/div[1]/div/div[1]/input'
song_id =  '//*[@id="mG61Hd"]/div[2]/div/div[2]/div[6]/div/div/div[2]/div/div[1]/div/div[1]/input'
field_ids = [userName_id,phrase_id,animal_id,talent_id,planet_id,song_id]

current_row = 1
for userData in sheet.iter_rows(min_row=1,max_row=6, values_only = True) :
    userName = userData[1]
    phrase = userData[2]
    animal = userData[3]
    talent = userData[4]
    planet = userData[5]
    song = userData[6]
    input_values = [userName,phrase,animal,talent,planet,song]
    print("row",input_values)

    for field_id, input_value in zip(field_ids, input_values):
        input_field=WebDriverWait(driver,4).until(
                EC.element_to_be_clickable((By.XPATH,field_id))
        )
        input_field.click()
        input_field.send_keys(input_value)
        time.sleep(0.5)
    try:
        submit_btn = driver.find_element(By.XPATH,'//*[@id="mG61Hd"]/div[2]/div/div[3]/div[1]/div[1]/div/span/span')
        submit_btn.click()
    except Exception as e:
        print(f"Not able to submit ${userData[1]} , ${e}")
    try:
        if current_row == sheet.max_row:
            print("all data successfully submited")
        else:
            WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/div[2]/div[1]/div/div[4]/a'))
            )

            driver.back()
            WebDriverWait(driver,4).until(
                EC.element_to_be_clickable((By.XPATH,'//*[@id="mG61Hd"]/div[2]/div/div[3]/div[1]/div[1]/div/span/span')))
            driver.refresh()
            WebDriverWait(driver,4).until(
                EC.element_to_be_clickable((By.XPATH,'//*[@id="mG61Hd"]/div[2]/div/div[3]/div[1]/div[1]/div/span/span')))
            print(current_row, f" ${userData[1]} successfully pass")
            current_row+=1
    except Exception as e:
        print(f"Not able to Run next Dataset next_row: ${current_row+1} max_row: ${sheet.max_row} " )     

print(f"last success data {userData[0]} ladke ka naam {userData[1]} iska data successful ho gya ")
wb.close()





